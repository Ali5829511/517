#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


def create_excel_report(report_type="all"):
    """إنشاء تقرير Excel احترافي"""

    # الاتصال بقاعدة البيانات
    conn = sqlite3.connect("/home/ubuntu/housing-system-deployment/housing_database.db")
    cursor = conn.cursor()

    # إنشاء workbook جديد
    wb = Workbook()

    # حذف الورقة الافتراضية
    wb.remove(wb.active)

    # إنشاء ورقة الغلاف
    cover_sheet = wb.create_sheet("الغلاف", 0)
    cover_sheet["A1"] = "تقرير نظام إدارة الإسكان الجامعي"
    cover_sheet["A1"].font = Font(size=24, bold=True, color="FFFFFF")
    cover_sheet["A1"].fill = PatternFill(
        start_color="667EEA", end_color="667EEA", fill_type="solid"
    )
    cover_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    cover_sheet.merge_cells("A1:F3")

    cover_sheet["A5"] = "جامعة الإمام محمد بن سعود الإسلامية"
    cover_sheet["A5"].font = Font(size=16, bold=True)
    cover_sheet["A5"].alignment = Alignment(horizontal="center")
    cover_sheet.merge_cells("A5:F5")

    cover_sheet["A7"] = f'تاريخ التقرير: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    cover_sheet["A7"].alignment = Alignment(horizontal="center")
    cover_sheet.merge_cells("A7:F7")

    # 1. تقرير الوحدات الشاغرة
    if report_type in ["all", "vacant_units"]:
        ws_vacant = wb.create_sheet("الوحدات الشاغرة")
        cursor.execute(
            """
            SELECT b.building_number, u.unit_number, u.unit_type, 'N/A' as floor, u.status
            FROM units u
            LEFT JOIN buildings b ON u.building_id = b.id
            WHERE u.status = 'شاغر' OR u.status = 'vacant'
            ORDER BY b.building_number, u.unit_number
        """
        )

        headers = ["المبنى", "رقم الوحدة", "نوع الوحدة", "الطابق", "الحالة"]
        add_sheet_data(ws_vacant, "الوحدات السكنية الشاغرة", headers, cursor.fetchall())

    # 2. تقرير السكان
    if report_type in ["all", "residents"]:
        ws_residents = wb.create_sheet("السكان")
        cursor.execute(
            """
            SELECT name, building, unit_number, phone, email, status
            FROM residents
            ORDER BY building, unit_number
        """
        )

        headers = ["الاسم", "المبنى", "رقم الوحدة", "الهاتف", "البريد الإلكتروني", "الحالة"]
        add_sheet_data(ws_residents, "قائمة السكان", headers, cursor.fetchall())

    # 3. تقرير الملصقات
    if report_type in ["all", "stickers"]:
        ws_stickers = wb.create_sheet("الملصقات")
        cursor.execute(
            """
            SELECT stickerNumber, plateNumber, ownerName, vehicleType, status
            FROM stickers
            ORDER BY stickerNumber
        """
        )

        headers = ["رقم الملصق", "رقم اللوحة", "اسم المالك", "نوع السيارة", "الحالة"]
        add_sheet_data(ws_stickers, "ملصقات السيارات", headers, cursor.fetchall())

    # 4. تقرير المواقف
    if report_type in ["all", "parking"]:
        ws_parking = wb.create_sheet("المواقف")
        cursor.execute(
            """
            SELECT id, plateNumber, building, type, status
            FROM parking
            ORDER BY building, id
        """
        )

        headers = ["رقم الموقف", "رقم اللوحة", "المبنى", "النوع", "الحالة"]
        add_sheet_data(ws_parking, "المواقف", headers, cursor.fetchall())

    # 5. تقرير المباني
    if report_type in ["all", "buildings"]:
        ws_buildings = wb.create_sheet("المباني")
        cursor.execute(
            """
            SELECT b.building_number, COUNT(*) as total_units,
                   SUM(CASE WHEN u.status = 'مشغول' OR u.status = 'occupied' THEN 1 ELSE 0 END) as occupied,
                   SUM(CASE WHEN u.status = 'شاغر' OR u.status = 'vacant' THEN 1 ELSE 0 END) as vacant
            FROM units u
            LEFT JOIN buildings b ON u.building_id = b.id
            GROUP BY b.building_number
            ORDER BY b.building_number
        """
        )

        headers = ["المبنى", "إجمالي الوحدات", "المشغولة", "الشاغرة"]
        add_sheet_data(ws_buildings, "إحصائيات المباني", headers, cursor.fetchall())

    conn.close()

    # حفظ الملف
    filename = f'/home/ubuntu/housing-system-deployment/static/reports/تقرير_شامل_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    return filename


def add_sheet_data(ws, title, headers, data):
    """إضافة بيانات إلى ورقة Excel مع تنسيق احترافي"""

    # العنوان
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A1:{chr(64 + len(headers))}2")

    # الرؤوس
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    # البيانات
    for row_idx, row_data in enumerate(data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # تلوين الصفوف بالتناوب
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # ضبط عرض الأعمدة
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20


if __name__ == "__main__":
    import sys

    report_type = sys.argv[1] if len(sys.argv) > 1 else "all"
    filename = create_excel_report(report_type)
    print(filename)
